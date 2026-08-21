# -*- coding: utf-8 -*-
"""
CKD 컨테이너 산출 자동화 — 샘플 엑셀 데이터 생성 스크립트

생성 파일:
  sample_data/plan_브라질.xlsx  : 브라질법인 18개월 생산계획 (모델 × 월)
  sample_data/plan_인도.xlsx    : 인도법인 18개월 생산계획
  sample_data/model_box.xlsx   : 모델별 CKD 박스 구성 (대당 박스 수량)
  sample_data/box_master.xlsx  : 박스 규격 마스터 (치수/부피/중량/적층/혼적 그룹)
  js/sample-data.js            : 위와 동일한 데이터의 임베디드 JSON
                                 (웹 UI의 "샘플 데이터 불러오기" 버튼용)

실행: python3 make_samples.py
"""

import json
import math
import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
random.seed(20260821)  # 항상 같은 샘플이 나오도록 시드 고정

# ---------------------------------------------------------------------------
# 1) 기준 정보 정의
# ---------------------------------------------------------------------------

MONTHS = []
_y, _m = 2026, 9
for _ in range(18):  # 2026-09 ~ 2028-02
    MONTHS.append(f"{_y}-{_m:02d}")
    _m += 1
    if _m > 12:
        _m, _y = 1, _y + 1

# 박스 마스터: (박스코드, 길이mm, 폭mm, 높이mm, 중량kg, 적층 가능 단수, 혼적 그룹, 비고)
# 부피(CBM)는 치수에서 자동 계산 (L×W×H / 1e9)
BOXES = [
    ("BX-A", 2300, 2200, 2200, 3200, 1, "G1", "메인프레임/하부주행체"),
    ("BX-B", 2250, 1100, 1100, 1800, 2, "G1", "붐/암 용접구조물"),
    ("BX-C", 1500, 1200, 1300, 1500, 2, "G1", "엔진/유압 유닛"),
    ("BX-D", 1150, 1150, 1150, 700, 3, "G2", "중형 부품(기어/실린더)"),
    ("BX-E", 1100, 900, 900, 350, 4, "G2", "소형 부품(볼트/베어링)"),
    ("BX-F", 800, 600, 700, 120, 5, "G2", "전장/호스류"),
    ("BX-G", 1900, 1600, 1900, 900, 1, "G2", "캐빈/외장 판넬"),
]

def cbm(l, w, h):
    return round(l * w * h / 1e9, 3)

# 모델-박스 구성: 모델 1대당 박스 종류별 수량
MODEL_BOX = {
    # 브라질법인 (중형/대형 굴착기 CKD)
    "R210-BR": {"BX-A": 1, "BX-B": 1, "BX-C": 1, "BX-D": 2, "BX-E": 2, "BX-F": 3, "BX-G": 1},
    "R300-BR": {"BX-A": 1, "BX-B": 1, "BX-C": 2, "BX-D": 3, "BX-E": 2, "BX-F": 4, "BX-G": 1},
    "R140-BR": {"BX-B": 1, "BX-C": 1, "BX-D": 2, "BX-E": 1, "BX-F": 2, "BX-G": 1},
    # 인도법인 (소형~대형 굴착기 CKD)
    "R215-IN": {"BX-A": 1, "BX-B": 1, "BX-C": 1, "BX-D": 2, "BX-E": 2, "BX-F": 3, "BX-G": 1},
    "R80-IN":  {"BX-C": 1, "BX-D": 1, "BX-E": 1, "BX-F": 2},
    "R330-IN": {"BX-A": 1, "BX-B": 2, "BX-C": 2, "BX-D": 3, "BX-E": 3, "BX-F": 4, "BX-G": 1},
}

CORP_MODELS = {
    "브라질": ["R210-BR", "R300-BR", "R140-BR"],
    "인도": ["R215-IN", "R80-IN", "R330-IN"],
}

# 모델별 월 기본 생산량(대) — 계절성/성장 추세를 반영해 변동을 준다
BASE_QTY = {
    "R210-BR": 34, "R300-BR": 18, "R140-BR": 26,
    "R215-IN": 46, "R80-IN": 70, "R330-IN": 14,
}

def gen_plan(model, month_index):
    """월별 생산 대수: 기본량 × (완만한 성장 + 계절성) + 소폭 랜덤."""
    base = BASE_QTY[model]
    growth = 1.0 + 0.012 * month_index                       # 18개월 완만한 성장
    season = 1.0 + 0.15 * math.sin(2 * math.pi * (month_index + 3) / 12.0)  # 계절성 ±15%
    noise = random.uniform(0.88, 1.12)
    return max(0, int(round(base * growth * season * noise)))

PLAN = {
    corp: {model: [gen_plan(model, i) for i in range(18)] for model in models}
    for corp, models in CORP_MODELS.items()
}

# ---------------------------------------------------------------------------
# 2) 엑셀 파일 생성 (공통 스타일)
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_plan(corp, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "생산계획"
    header = ["법인", "모델명"] + MONTHS
    ws.append(header)
    for model in CORP_MODELS[corp]:
        ws.append([corp, model] + PLAN[corp][model])
    style_header(ws, len(header))
    autofit(ws, [10, 12] + [10] * len(MONTHS))
    ws.freeze_panes = "C2"
    wb.save(path)


def write_model_box(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "모델-박스 구성"
    header = ["모델명", "박스코드", "대당 박스 수량"]
    ws.append(header)
    for model, boxes in MODEL_BOX.items():
        for code, qty in boxes.items():
            ws.append([model, code, qty])
    style_header(ws, len(header))
    autofit(ws, [12, 12, 16])
    ws.freeze_panes = "A2"
    wb.save(path)


def write_box_master(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "박스 마스터"
    header = ["박스코드", "길이(mm)", "폭(mm)", "높이(mm)", "부피(CBM)",
              "중량(kg)", "적층 가능 단수", "혼적 그룹", "비고"]
    ws.append(header)
    for code, l, w, h, kg, stack, group, note in BOXES:
        ws.append([code, l, w, h, cbm(l, w, h), kg, stack, group, note])
    style_header(ws, len(header))
    autofit(ws, [10, 10, 10, 10, 11, 10, 14, 10, 24])
    ws.freeze_panes = "A2"
    wb.save(path)


# ---------------------------------------------------------------------------
# 3) 웹 UI용 임베디드 JSON (js/sample-data.js) — 엑셀과 항상 동일 데이터
# ---------------------------------------------------------------------------

def write_sample_js(path):
    plan_rows = []
    for corp, models in CORP_MODELS.items():
        for model in models:
            plan_rows.append({
                "corp": corp,
                "model": model,
                "qty": {mo: q for mo, q in zip(MONTHS, PLAN[corp][model])},
            })
    model_box_rows = [
        {"model": model, "boxCode": code, "perUnit": qty}
        for model, boxes in MODEL_BOX.items()
        for code, qty in boxes.items()
    ]
    box_master = {
        code: {
            "lengthMm": l, "widthMm": w, "heightMm": h,
            "cbm": cbm(l, w, h), "weightKg": kg,
            "maxStack": stack, "group": group, "note": note,
        }
        for code, l, w, h, kg, stack, group, note in BOXES
    }
    data = {
        "months": MONTHS,
        "plan": plan_rows,
        "modelBox": model_box_rows,
        "boxMaster": box_master,
    }
    body = json.dumps(data, ensure_ascii=False, indent=2)
    js = (
        "/**\n"
        " * 샘플 데이터 (make_samples.py가 생성 — 직접 수정하지 말 것)\n"
        " * sample_data/*.xlsx 와 항상 동일한 내용이다.\n"
        " */\n"
        "(function (root, factory) {\n"
        "  if (typeof module !== 'undefined' && module.exports) {\n"
        "    module.exports = factory();\n"
        "  } else {\n"
        "    root.CKD_SAMPLE_DATA = factory();\n"
        "  }\n"
        "})(typeof self !== 'undefined' ? self : this, function () {\n"
        "  'use strict';\n"
        "  return " + body + ";\n"
        "});\n"
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(js)


def main():
    out_dir = os.path.join(BASE, "sample_data")
    os.makedirs(out_dir, exist_ok=True)

    write_plan("브라질", os.path.join(out_dir, "plan_브라질.xlsx"))
    write_plan("인도", os.path.join(out_dir, "plan_인도.xlsx"))
    write_model_box(os.path.join(out_dir, "model_box.xlsx"))
    write_box_master(os.path.join(out_dir, "box_master.xlsx"))
    write_sample_js(os.path.join(BASE, "js", "sample-data.js"))

    total = {c: sum(sum(v) for v in PLAN[c].values()) for c in PLAN}
    print("샘플 데이터 생성 완료")
    print(f"  기간: {MONTHS[0]} ~ {MONTHS[-1]} ({len(MONTHS)}개월)")
    for corp, tot in total.items():
        print(f"  {corp}법인 총 생산계획: {tot:,}대")
    print("  파일: sample_data/*.xlsx, js/sample-data.js")


if __name__ == "__main__":
    main()
